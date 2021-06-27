import os
from dataclasses import dataclass
from typing import List
from flask import make_response
import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from src.service.image_process import Word
from shapely.geometry.point import Point


@dataclass(frozen=True)
class ExcelData:
    text: str
    x: int  # ExcelのX方向の座標
    y: int  # Excelのy方向の座標


class ExcelProcess:
    """ Excel処理関係
    """

    def __init__(self):
        """ コンストラクタ
        @param aBlocks [Block]
        @param aWords [Word]
        """

        # # ブロックの配列が入った配列
        # self.blocks: List[Block] = []
        # # 文字とその座標が入った配列
        # self.words: List[Word] = []
        # Exel変換後のデータ配列
        # [{"text":???, "x":???, "y":???}, ... ]
        # self.outputExcelData: List[ExcelData] = []

        # if len(aBlocks) > 0 and len(aWords) > 0:
        #     self.blocks = aBlocks
        #     self.words = aWords

    def mappingFromPixelCoordinatesToExcelCoordinates(self, words: List[Word]) -> List[ExcelData]:
        """
        画像の座標()から
        エクセルの座標({"text":???,"x":??,"y":??})に変換する
        """
        outputExcelData: List[ExcelData] = []
        # Excel1セルあたりのピクセル数
        oneExcelCellRate: float = self.__getAverageCharacterSizeFromPixcelCoordinates(words)
        for word in words:
            # ポリゴンの重心をエクセル方眼紙にマッピング
            outputExcelData.append(
                ExcelData(
                    text=word.text,
                    x=int(word.poly.centroid.x / oneExcelCellRate),
                    y=int(word.poly.centroid.y / oneExcelCellRate)
                )
            )
        return outputExcelData

    def __getAverageCharacterSizeFromPixcelCoordinates(self, words: List[Word]) -> float:
        """
        テキストから[{"text":???, "id", "confidence":???, poly:Polygon([(xy)(xy)(xy)(xy)])}, ... ]
        平均文字サイズを取得する（ピクセル）
        """
        totalDistance: float = 0
        for word in words:
            totalDistance = totalDistance + Point(word.poly.exterior.xy[0][0],
                                                  word.poly.exterior.xy[1][0]).distance(Point(word.poly.exterior.xy[0][3],
                                                                                              word.poly.exterior.xy[1][3]))
        return float(totalDistance) / float(len(words))

    def sampleExcel(self):
        # ファイルを開く
        wb: Workbook = openpyxl.load_workbook("secret/excelTemplate/Sample.xlsx")
        # シート指定
        ws: Worksheet = wb["Sheet1"]
        # セル指定
        c1: Cell = ws.cell(row=1, column=1)
        # セルへ書き込み
        c1.value = "test"
        wb.save("output.xlsx")

    def writeExcel(self, outputExcelData: List[ExcelData]):
        """ エクセルファイルに書き込む
        """
        templateFilePath: str = "secret/excelTemplate/Sample.xlsx"
        outputFilePath: str = "secret/excelTemplate/output.xlsx"
        wb: Workbook = openpyxl.load_workbook(templateFilePath)
        ws: Worksheet = wb["Sheet1"]
        for data in outputExcelData:
            c1: Cell = ws.cell(row=data.y, column=data.x)
            c1.value = data.text
        wb.save(outputFilePath)

    def getResponse(self):
        """ レスポンスを作成する
        """
        # 出力Excelファイルのパス
        outputExcelFileName = "output.xlsx"
        outputExcelFilePath = "secret/excelTemplate/" + outputExcelFileName
        # Flaskのレスポンスを作成する
        response = make_response()
        # ダウンロードするEXCELファイルを開く
        wb = open(outputExcelFilePath, "rb")
        # ダウンロードするEXCELファイルをレスポンスに設定する
        response.data = wb.read()
        # ダウンロードするEXCELファイルをクローズ
        wb.close()
        # ダウンロードするEXCELファイル名を設定
        response.headers["Content-Disposition"] = "attachment; filename=" + outputExcelFileName
        # MIMEタイプをレスポンスに設定
        response.mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        # MIMEタイプをレスポンスに設定
        os.remove(outputExcelFilePath)
        return response
